"""
Author Linnea Shieh <laiello@stanford.edu>
Copyright Stanford University 2019

Script for adding serial title information to Springer COUNTER book reports.
"""

import argparse
import logging
import openpyxl
import re
import requests
from lxml import html


VERSION = 1

 
def ParseCommandArgs():
  """
  Parses argument information passed on command line, which contains user options.

  Returns:
    A namespace called args populated with values for arguments.
  """
  parser = argparse.ArgumentParser(
      description="Add serial info to Springer book reports.")
  parser.add_argument("first_row", type=int, default=10, nargs="?", 
		      help="From XLS, row number for first book.") 
  parser.add_argument("doi_column", type=str, default="D", nargs="?",
		      help="From XLS, column letter for the DOI.") 
  parser.add_argument("issn_column", type=str, default="G", nargs="?",
		      help="From XLS, column letter for the ISSN.") 
  parser.add_argument("filename", type=str, nargs="?",
   	   	      default="Springer2018BookReport3.xlsx",
		      help="Filename of Springer book report.")
  parser.add_argument("--rerun", dest="rerun", action="store_true",
                      help="Resume a previous run.")
  args = parser.parse_args()
  return args


def RunReportEnhancement(filename, first_row, doi_col, issn_col, rerun):
  """
  Adds new columns and enhanced report with publishing and series metadata.

  Args:
    filename: (str) Path (relative or absolute) to the input Excel file.
    first_row: (int) Excel row representing the first book (no headers).
    doi_col: (str) Column letter containing the book DOI.
    issn_col: (str) Column letter containing the book ISSN.
    rerun: (bool) Whether this is a rerun (i.e. output Excel already created).
  """
  book_count = 0
  split_filename = filename.split(".")
  new_filename = split_filename[0] + "_v" + str(VERSION) + "." + split_filename[1]

  series_col = ord(issn_col) - 63
  acronym_col = series_col + 1
  volume_col = series_col + 2
  year_col = series_col + 3
  package_col = series_col + 4
  subseries_col = series_col + 5
  if rerun:
    wb = openpyxl.load_workbook(new_filename)
    ws = wb["Sheet0"]
  else:
    wb = openpyxl.load_workbook(filename)
    ws = wb["Sheet0"]
    ws.insert_cols(series_col, amount=6)

  # for i in range(8960, 8980):
  for i in range(first_row, ws.max_row + 1):
    issn = ws.cell(row=i, column=(ord(issn_col) - 64)).value
    if issn:
      book_count += 1
      doi = ws.cell(row=i, column=(ord(doi_col) - 64)).value
      book_html, landolt = RequestBookInfoPage(doi)
      if landolt:
        book_dict = ParseLandoltBookPage(book_html)
      else:
        book_dict = ParseBookPage(book_html)
      
      logging.debug("Row %d: %s", i, book_dict["series"])
      ws.cell(row=i, column=series_col).value = book_dict["series"]
      ws.cell(row=i, column=acronym_col).value = book_dict["acronym"]
      ws.cell(row=i, column=volume_col).value = book_dict["volume"]
      ws.cell(row=i, column=year_col).value = book_dict["year"]
      ws.cell(row=i, column=package_col).value = book_dict["package"]
      ws.cell(row=i, column=subseries_col).value = book_dict["subseries"]
    if i % 1000 == 0: wb.save(new_filename)
  
  wb.save(new_filename)
  logging.info("Enhanced %d books.", book_count)


def RequestBookInfoPage(doi):
  """
  Retrieves HTML content of a book's information page at link.springer.com.

  Args:
    doi: (str) DOI of book.
  
  Returns:
    An HTML document representing the book's information webpage, and a bool
    indicating that the book is part of the Landolt-Bornstein series.
  """
  base_url = ("http://link.springer.com/")
  landolt = False

  # The book info page is sought for in 3 places in sequence. Resorting to the
  # dx.doi.org URL means the book is in the Landolt-Bornstein series.
  book_request = base_url + "book/" + doi + "#about" 
  book_page = requests.get(book_request)        

  if book_page.status_code == 404:
    book_request = base_url + "referencework/" + doi + "#about"
    book_page = requests.get(book_request)
       
    if book_page.status_code == 404:
      book_request = "http://dx.doi.org/" + doi
      book_page = requests.get(book_request)
      landolt = True      

      if book_page.status_code == 404:
        return None, False

  return html.fromstring(book_page.content), landolt


def ParseBookPage(book_html):
  """
  Parses HTML of the book's info page and stores desired metadata.

  Args:
    book_html: (HTML Element) HTML representing the book's info page.

  Returns:
    A dict populated with series and publishing metadata.
  """
  book_dict = dict(series="Unavailable", acronym="", volume="",
                   year="", package="", subseries="")
  if book_html is not None:
    series_text = book_html.xpath("//p[@data-test='test-series']/a/text()")
    if series_text: book_dict["series"] = str(series_text[0].encode("utf-8"))
    volume_text = book_html.xpath("//p[@data-test='test-series']/span/text()")
    if volume_text:
      a = re.search(r"\(([A-Za-z\+]+)", str(volume_text[0].encode("utf-8")))
      if a: book_dict["acronym"] = a.group(1)
      v = re.search(r"volume (.+)\)", str(volume_text[0].encode("utf-8")))
      if v: book_dict["volume"] = v.group(1)
    year_text = book_html.xpath("//span[@id='copyright-info']/text()")
    if year_text:
      y = re.search(r"\d\d\d\d", str(year_text[0].encode("utf-8")))
      if y: book_dict["year"] = int(y.group(0))
    package_text = book_html.xpath("//a[@id='ebook-package']/text()")
    if package_text: book_dict["package"] = str(package_text[0].encode("utf-8"))
    subseries_text = book_html.xpath("//p[@data-test='test-subseries']/a/text()")
    if subseries_text: book_dict["subseries"] = str(subseries_text[0].encode("utf-8"))
  
  return book_dict


def ParseLandoltBookPage(book_html):
  """
  Parses HTML of a Landolt-Bornstein book's info page and stores metadata.

  Args:
    book_html: (HTML Element) HTML representing the book's info page.

  Returns:
    A dict populated with series and publishing metadata.
  """
  book_dict = dict(series="Unavailable", acronym="", volume="",
                   year="", package="", subseries="")
  if book_html is not None:
    series_text = book_html.xpath(
        "//div[@class='publication-title']/span/text()")
    if series_text: book_dict["series"] = str(series_text[0].encode("utf-8"))
    volume_text = book_html.xpath(
        "//div[@class='document__enumeration']/span/text()")
    if volume_text:
      v = re.search(r"Volume (.+) ", str(volume_text[0].encode("utf-8")))
      if v: book_dict["volume"] = v.group(1)
      y = re.search(r"\d\d\d\d", str(volume_text[0].encode("utf-8")))
      if y: book_dict["year"] = int(y.group(0))
  
  return book_dict


def main():
  logging.basicConfig(format="%(asctime)s %(message)s", level=logging.DEBUG)
  args = ParseCommandArgs()
  RunReportEnhancement(args.filename, args.first_row, args.doi_column,
                       args.issn_column, args.rerun)


if __name__ == "__main__":
  main()
